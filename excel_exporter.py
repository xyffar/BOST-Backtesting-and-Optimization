# excel_exporter.py

import datetime as dt
import os
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
import xlsxwriter
from openpyxl.styles import Font
from openpyxl.utils import (
    get_column_letter,
)  # Per convertire numeri di colonna in lettere

from config import (
    HEADER_COLORS,
    HEADER_FONT,  # Import MESSAGES["display_texts"]
    MESSAGES,
    ROW_STRIPE_COLORS,
    THIN_BORDER,
)


def _write_comparison_table(
    worksheet: xlsxwriter.worksheet.Worksheet,
    workbook: xlsxwriter.workbook.Workbook,
    comparison_df: pd.DataFrame,
    start_row: int,
    start_col: int = 0,
) -> int:
    """Write the comparison DataFrame (strategy vs. benchmark) as an Excel table.

    This function formats the data, handles NaN/Inf values, and applies table formatting
    with headers and data rows.

    Args:
        worksheet (xlsxwriter.worksheet.Worksheet): The XlsxWriter worksheet object to write to.
        workbook (xlsxwriter.workbook.Workbook): The XlsxWriter workbook object.
        comparison_df (pd.DataFrame): DataFrame containing the comparison statistics.
        start_row (int): The starting row for writing the table.
        start_col (int, optional): The starting column for writing the table. Defaults to 0.

    Returns:
        int: The next available row after writing the table.

    """
    if comparison_df is None or comparison_df.empty:
        worksheet.write_string(
            start_row,
            start_col,
            MESSAGES["display_texts"]["excel_exporter"]["no_comparison_data"],
            workbook.add_format({"italic": True}),
        )
        return start_row + 2

    # Prepara il DataFrame per la tabella Excel
    # Assicurati che l'indice sia una colonna normale per la tabella
    df_to_export = comparison_df.reset_index()
    df_to_export.rename(
        columns={"index": MESSAGES["display_texts"]["excel_exporter"]["metric_column_name"]},
        inplace=True,
    )

    # Arrotonda i valori numerici a 2 decimali per la visualizzazione
    for col in df_to_export.select_dtypes(include=np.number).columns:
        df_to_export[col] = df_to_export[col].round(2)

    # Sostituisci i valori NaN e Inf
    df_to_export = df_to_export.replace([np.inf, -np.inf], MESSAGES["display_texts"]["excel_exporter"]["inf_value"])
    df_to_export = df_to_export.fillna(MESSAGES["display_texts"]["excel_exporter"]["nan_value"])

    # Titolo della sezione
    worksheet.write_string(
        start_row,
        start_col,
        MESSAGES["display_texts"]["excel_exporter"]["comparison_table_title"],
        workbook.add_format({"bold": True, "font_size": 12}),
    )
    current_row = start_row + 2  # Spazio per il titolo

    # Scrivi gli header
    header_format = workbook.add_format({"bold": True, "bg_color": "#D7E4BC", "border": 1})
    for col_num, value in enumerate(df_to_export.columns.values):
        worksheet.write(current_row, start_col + col_num, value, header_format)

    # Scrivi i dati
    for row_num, row_data in df_to_export.iterrows():
        for col_num, cell_value in enumerate(row_data.values):
            worksheet.write(current_row + 1 + row_num, start_col + col_num, cell_value)

    # Definisci l'intervallo per la tabella Excel
    end_row = current_row + df_to_export.shape[0]
    end_col = start_col + df_to_export.shape[1] - 1

    # Aggiungi la formattazione tabella
    table_range = xlsxwriter.utility.xl_range(current_row, start_col, end_row, end_col)
    worksheet.add_table(
        table_range,
        {
            "data": df_to_export.values.tolist(),
            "columns": [{"header": col} for col in df_to_export.columns],
        },
    )

    return end_row + 2  # Restituisce la riga successiva disponibile


def _write_trades_table(
    worksheet: xlsxwriter.worksheet.Worksheet,
    workbook: xlsxwriter.workbook.Workbook,
    trades_df: pd.DataFrame,
    start_row: int,
    start_col: int = 0,
) -> int:
    """Write the trades DataFrame as an Excel table.

    This function formats the trade details, handles NaN/Inf values, and applies
    table formatting to the data.

    Args:
        worksheet (xlsxwriter.worksheet.Worksheet): The XlsxWriter worksheet object to write to.
        workbook (xlsxwriter.workbook.Workbook): The XlsxWriter workbook object.
        trades_df (pd.DataFrame): DataFrame containing the trade details.
        start_row (int): The starting row for writing the table.
        start_col (int, optional): The starting column for writing the table. Defaults to 0.

    Returns:
        int: The next available row after writing the table.

    """
    if trades_df is None or trades_df.empty:
        worksheet.write_string(
            start_row,
            start_col,
            MESSAGES["display_texts"]["excel_exporter"]["no_trades_executed_excel"],
            workbook.add_format({"italic": True}),
        )
        return start_row + 2

    # Arrotonda i valori numerici a 2 decimali per la visualizzazione
    df_to_export = trades_df.copy()
    for col in df_to_export.select_dtypes(include=np.number).columns:
        df_to_export[col] = df_to_export[col].round(2)

    # Sostituisci i valori NaN e Inf
    df_to_export = df_to_export.replace([np.inf, -np.inf], MESSAGES["display_texts"]["excel_exporter"]["inf_value"])
    df_to_export = df_to_export.fillna(MESSAGES["display_texts"]["excel_exporter"]["nan_value"])

    # Titolo della sezione
    worksheet.write_string(
        start_row,
        start_col,
        MESSAGES["display_texts"]["excel_exporter"]["trades_table_title"],
        workbook.add_format({"bold": True, "font_size": 12}),
    )
    current_row = start_row + 2  # Spazio per il titolo

    # Scrivi gli header
    header_format = workbook.add_format({"bold": True, "bg_color": "#D7E4BC", "border": 1})
    for col_num, value in enumerate(df_to_export.columns.values):
        worksheet.write(current_row, start_col + col_num, value, header_format)

    # Scrivi i dati
    for row_num, row_data in df_to_export.iterrows():
        for col_num, cell_value in enumerate(row_data.values):
            # Gestione speciale per date/timestamp se necessario, altrimenti scrivi il valore
            if isinstance(cell_value, (dt.datetime, pd.Timestamp)):
                worksheet.write_string(current_row + 1 + row_num, start_col + col_num, str(cell_value))
            else:
                worksheet.write(current_row + 1 + row_num, start_col + col_num, cell_value)

    # Definisci l'intervallo per la tabella Excel
    end_row = current_row + df_to_export.shape[0]
    end_col = start_col + df_to_export.shape[1] - 1

    # Aggiungi la formattazione tabella
    table_range = xlsxwriter.utility.xl_range(current_row, start_col, end_row, end_col)
    worksheet.add_table(
        table_range,
        {
            "data": df_to_export.values.tolist(),
            "columns": [{"header": col} for col in df_to_export.columns],
        },
    )

    return end_row + 2  # Restituisce la riga successiva disponibile


def export_to_excel(
    all_ticker_results: dict[str, tuple[pd.Series, pd.DataFrame, pd.DataFrame]],
    filename: str = "report_backtest_strategie.xlsx",  # Questo parametro verrà ora sovrascritto
) -> BytesIO:
    """Export backtest results and trade details to an Excel file with separate sheets per ticker.

    Each sheet includes two tables: comparison statistics and trade details.
    The function generates an in-memory Excel file using XlsxWriter.

    Args:
        all_ticker_results (dict[str, tuple[pd.Series, pd.DataFrame, pd.DataFrame]]):
            A dictionary where keys are ticker symbols and values are tuples containing
            (strategy_stats, trades_df, benchmark_comparison_df).
        filename (str, optional): The base name for the Excel file. Note that a timestamp
                                  is often appended to create a dynamic filename,
                                  though in this current implementation it's not used directly
                                  for the in-memory BytesIO object. Defaults to "report_backtest_strategie.xlsx".

    Returns:
        BytesIO: An in-memory BytesIO object containing the generated Excel file.

    """
    output: BytesIO = BytesIO()

    # Genera il nome del file con timestamp
    # timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    # dynamic_filename = f"BacktestResult_{timestamp}.xlsx"

    # Usa XlsxWriter come motore per pandas per più opzioni di formattazione
    # È più semplice creare un Workbook diretto e poi aggiungere i fogli e le tabelle
    workbook: xlsxwriter.workbook.Workbook = xlsxwriter.Workbook(output, {"in_memory": True})

    if not all_ticker_results:
        # Crea un foglio vuoto con un messaggio se non ci sono risultati
        worksheet = workbook.add_worksheet(MESSAGES["display_texts"]["excel_exporter"]["no_results_sheet_name"])
        worksheet.write_string(0, 0, MESSAGES["display_texts"]["excel_exporter"]["no_results_found"])
    else:
        for ticker, (
            trades_df,
            benchmark_comparison_df,
        ) in all_ticker_results.items():
            # Pulisci il nome del ticker per usarlo come nome del foglio (max 31 caratteri)
            sheet_name = ticker.replace(":", "_").replace("/", "_")[:31]
            worksheet = workbook.add_worksheet(sheet_name)

            # Imposta larghezza colonne per leggibilità
            worksheet.set_column(0, 5, 20)  # Colonna 0 a 5, larghezza 20

            current_row = 0

            # Scrivi la tabella di confronto
            current_row = _write_comparison_table(worksheet, workbook, benchmark_comparison_df, current_row)

            # Aggiungi uno spazio tra le due tabelle
            current_row += 2

            # Scrivi la tabella dei trade
            current_row = _write_trades_table(worksheet, workbook, trades_df, current_row)

    workbook.close()  # Chiudi la cartella di lavoro per salvare i dati

    output.seek(0)
    return output


def get_next_empty_row_in_column(ws: openpyxl.worksheet.worksheet.Worksheet, col_idx: int, start_row: int) -> int:
    """Find the next empty row in a specific column of an Openpyxl worksheet, starting from `start_row`.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The Openpyxl worksheet object.
        col_idx (int): The 1-based index of the column to check for empty rows.
        start_row (int): The 1-based row number from which to start the search.

    Returns:
        int: The 1-based row number of the next empty row found.

    """
    row = start_row
    while ws.cell(row=row, column=col_idx).value is not None:
        row += 1
    return row


def _read_sheet_metadata(wb: openpyxl.workbook.workbook.Workbook, metadata_sheet_name: str) -> dict:
    """Read sheet layout metadata from a hidden sheet within an Openpyxl workbook.

    This metadata includes information like header color indices for different sheets.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The Openpyxl workbook object.
        metadata_sheet_name (str): The name of the hidden sheet where metadata is stored.

    Returns:
        dict: A dictionary mapping sheet names to their metadata (e.g., {"header_color_idx": int}).

    """
    metadata = {}
    if metadata_sheet_name in wb.sheetnames:
        ws_meta = wb[metadata_sheet_name]
        for row in range(2, ws_meta.max_row + 1):  # Salta la riga dell'intestazione
            sheet_name = ws_meta.cell(row=row, column=1).value
            color_idx = ws_meta.cell(row=row, column=2).value
            if sheet_name and color_idx is not None:
                metadata[sheet_name] = {"header_color_idx": int(color_idx)}
    return metadata


def _write_sheet_metadata(wb: openpyxl.workbook.workbook.Workbook, metadata_sheet_name: str, metadata: dict) -> None:
    """Write sheet layout metadata to a hidden sheet within an Openpyxl workbook.

    Clear existing content in the metadata sheet and then write the provided metadata, ensuring the sheet remains hidden.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The Openpyxl workbook object.
        metadata_sheet_name (str): The name of the hidden sheet where metadata will be stored.
        metadata (dict): A dictionary containing sheet metadata to be written
                         (e.g., {"Sheet_Name": {"header_color_idx": color_index}}).

    Returns:
        None

    """
    if metadata_sheet_name not in wb.sheetnames:
        ws_meta = wb.create_sheet(metadata_sheet_name)
    else:
        ws_meta = wb[metadata_sheet_name]

    ws_meta.sheet_state = "hidden"  # Assicurati che sia nascosto

    # Cancella il contenuto esistente (semplice riscrittura)
    for row in range(1, ws_meta.max_row + 1):
        for col in range(1, ws_meta.max_column + 1):
            ws_meta.cell(row=row, column=col, value=None)

    ws_meta.cell(row=1, column=1, value="Sheet_Name")
    ws_meta.cell(row=1, column=2, value="Header_Color_Index")

    for row_idx, (sheet_name, data) in enumerate(metadata.items(), start=2):
        ws_meta.cell(row=row_idx, column=1, value=sheet_name)
        ws_meta.cell(row=row_idx, column=2, value=data["header_color_idx"])
    # if metadata_sheet_name not in wb.sheetnames:
    #     ws_meta = wb.create_sheet(metadata_sheet_name)
    # else:
    #     ws_meta = wb[metadata_sheet_name]

    # ws_meta.sheet_state = "hidden"  # Assicurati che sia nascosto

    # # Cancella il contenuto esistente (semplice riscrittura)
    # for row in range(1, ws_meta.max_row + 1):
    #     for col in range(1, ws_meta.max_column + 1):
    #         ws_meta.cell(row=row, column=col, value=None)

    # ws_meta.cell(row=1, column=1, value="Sheet_Name")
    # ws_meta.cell(row=1, column=2, value="Header_Color_Index")

    # for row_idx, (sheet_name, data) in enumerate(metadata.items(), start=2):
    #     ws_meta.cell(row=row_idx, column=1, value=sheet_name)
    #     ws_meta.cell(row=row_idx, column=2, value=data["header_color_idx"])


def _apply_header_style(cell: openpyxl.cell.cell.Cell, color_idx: int) -> None:
    """Apply predefined header styling (font, fill color, border) to an Openpyxl cell.

    Args:
        cell (openpyxl.cell.cell.Cell): The Openpyxl cell object to style.
        color_idx (int): The index to select a header color from `HEADER_COLORS`.

    Returns:
        None

    """
    cell.font = HEADER_FONT
    cell.fill = HEADER_COLORS[color_idx % len(HEADER_COLORS)]
    cell.border = THIN_BORDER


def _apply_data_style(cell: openpyxl.cell.cell.Cell, row_num: int, header_row: int) -> None:
    """Apply data styling (striped rows and border) to an Openpyxl cell based on its row number.

    Args:
        cell (openpyxl.cell.cell.Cell): The Openpyxl cell object to style.
        row_num (int): The 1-based row number of the cell.
        header_row (int): The 1-based row number where the header is located,
                          used to calculate the stripe pattern.

    Returns:
        None

    """
    cell.fill = ROW_STRIPE_COLORS[(row_num - (header_row + 1)) % len(ROW_STRIPE_COLORS)]
    cell.border = THIN_BORDER


def _set_column_widths(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list,
    start_col_idx: int,
    current_data_row: int,
    header_row: int,
) -> None:
    """Automatically adjust the width of columns in an Openpyxl worksheet based on the maximum content length within each column.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The Openpyxl worksheet object.
        headers (list): A list of header names for the columns.
        start_col_idx (int): The 1-based starting column index.
        current_data_row (int): The 1-based row number of the last data entry.
        header_row (int): The 1-based row number where the headers are located.

    Returns:
        None

    """
    for col_offset, header in enumerate(headers):
        col_letter = get_column_letter(start_col_idx + col_offset)
        # Calcola la lunghezza massima del contenuto nella colonna
        max_len = len(str(header))  # Inizia con la lunghezza dell'intestazione
        for r_idx in range(header_row + 1, current_data_row + 1):
            cell_value = ws.cell(row=r_idx, column=start_col_idx + col_offset).value
            if cell_value is not None:
                # Per i float, considera la formattazione visualizzata
                if isinstance(cell_value, float):
                    max_len = max(max_len, len(f"{cell_value:.6f}"))  # Usa la formattazione di default per float
                else:
                    max_len = max(max_len, len(str(cell_value)))

        # Imposta la larghezza della colonna con un po' di padding
        ws.column_dimensions[col_letter].width = max_len + 2


def log_execution_data(start_time: float, end_time: float, action: str | None = None, **pars_time_log: dict) -> None:
    """Log execution data, including duration, action, and additional parameters to an Excel file.

    It maintains a main log sheet and separate sheets for specific actions,
    applying consistent styling and dynamically adjusting column widths.
    The function uses Openpyxl to interact with the Excel file, ensuring
    existing data and formatting are preserved.

    Args:
        start_time (float): The start timestamp of the action (e.g., from `time.perf_counter()`).
        end_time (float): The end timestamp of the action (e.g., from `time.perf_counter()`).
        action (str, optional): A descriptive string for the action being logged
                                (e.g., "Backtest", "Optimization"). If None,
                                a specific action log sheet will not be created.
        **pars_time_log: Arbitrary keyword arguments representing additional parameters
                         or context for the logged action (e.g., 'periods', 'strategy').

    Returns:
        None: This function performs file I/O and prints messages directly.

    """
    excel_file = "time_logs.xlsx"
    main_sheet_name = "Log"
    metadata_sheet_name = "_Metadata"

    execution_duration = end_time - start_time
    current_timestamp = dt.datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    try:
        wb = _load_or_create_workbook(excel_file)
        sheet_layouts = _read_sheet_metadata(wb, metadata_sheet_name)
        ws_main, main_header_color_idx, header_row = _prepare_main_log_sheet(wb, main_sheet_name, sheet_layouts)
        current_main_log_row = _write_main_log_row(
            ws_main, main_header_color_idx, header_row, execution_duration, action, current_timestamp
        )
        _set_column_widths(ws_main, ["Duration [s]", "Action", "Timestamp"], 1, current_main_log_row, header_row)

        if action is None:
            print("AVVISO: 'action' non specificato. Nessun log specifico per azione verrà creato.")
            _write_sheet_metadata_and_save(wb, metadata_sheet_name, sheet_layouts, excel_file)
            return

        _write_action_log_sheet(
            wb,
            action,
            sheet_layouts,
            header_row,
            execution_duration,
            current_timestamp,
            pars_time_log,
        )
        _write_sheet_metadata_and_save(wb, metadata_sheet_name, sheet_layouts, excel_file)
    except Exception as e:
        print(f"Errore generale durante la registrazione e formattazione dei dati di esecuzione: {e}")


def _write_sheet_metadata_and_save(
    wb: openpyxl.workbook.workbook.Workbook, metadata_sheet_name: str, sheet_layouts: dict, excel_file: str
) -> None:
    """Write sheet metadata and save the workbook to disk.

    Args:
        wb (openpyxl.workbook.workbook.Workbook): The Openpyxl workbook object.
        metadata_sheet_name (str): The name of the hidden sheet for metadata.
        sheet_layouts (dict): The metadata dictionary to write.
        excel_file (str): The file path to save the workbook.

    Returns:
        None

    """
    _write_sheet_metadata(wb, metadata_sheet_name, sheet_layouts)
    wb.save(excel_file)
    print(f"Dati di log scritti e formattati in '{excel_file}'.")


def _load_or_create_workbook(excel_file: str) -> openpyxl.Workbook:
    if os.path.exists(excel_file):
        return openpyxl.load_workbook(excel_file)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _prepare_main_log_sheet(
    wb: openpyxl.workbook.workbook.Workbook, main_sheet_name: str, sheet_layouts: dict
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, int, int]:
    main_log_name = "Main Log Summary"
    main_log_key = "MAIN_LOG_SHEET"
    main_log_headers = ["Duration [s]", "Action", "Timestamp"]
    header_row = 2

    if main_sheet_name not in wb.sheetnames:
        ws_main = wb.create_sheet(main_sheet_name, 0)
    else:
        ws_main = wb[main_sheet_name]

    if main_log_key not in sheet_layouts:
        next_color_idx = len(sheet_layouts) % len(HEADER_COLORS)
        sheet_layouts[main_log_key] = {"header_color_idx": next_color_idx}
    main_header_color_idx = sheet_layouts[main_log_key]["header_color_idx"]

    title_cell_main = ws_main.cell(row=1, column=1, value=main_log_name)
    title_cell_main.font = Font(bold=True, size=12, color="000000")

    if ws_main.cell(row=header_row, column=1).value != main_log_headers[0]:
        for col_idx, header_name in enumerate(main_log_headers):
            cell = ws_main.cell(row=header_row, column=1 + col_idx, value=header_name)
            _apply_header_style(cell, main_header_color_idx)

    return ws_main, main_header_color_idx, header_row


def _write_main_log_row(
    ws_main: openpyxl.worksheet.worksheet.Worksheet,
    main_header_color_idx: int,  # Not accessed, but kept for signature compatibility
    header_row: int,
    execution_duration: float,
    action: str,
    current_timestamp: str,
) -> int:
    main_data_cells = [execution_duration, action, current_timestamp]
    current_main_log_row = get_next_empty_row_in_column(ws_main, 1, header_row + 1)
    for col_idx, value in enumerate(main_data_cells):
        cell = ws_main.cell(row=current_main_log_row, column=1 + col_idx, value=value)
        _apply_data_style(cell, current_main_log_row, header_row)
    return current_main_log_row


def _write_action_log_sheet(
    wb: openpyxl.workbook.workbook.Workbook,
    action: str,
    sheet_layouts: dict,
    header_row: int,
    execution_duration: float,
    current_timestamp: str,
    pars_time_log: dict,
) -> None:
    action_sheet_name = action
    if action_sheet_name not in wb.sheetnames:
        ws_action = wb.create_sheet(action_sheet_name)
    else:
        ws_action = wb[action_sheet_name]

    if action_sheet_name not in sheet_layouts:
        next_color_idx = len(sheet_layouts) % len(HEADER_COLORS)
        sheet_layouts[action_sheet_name] = {"header_color_idx": next_color_idx}
    action_header_color_idx = sheet_layouts[action_sheet_name]["header_color_idx"]

    current_call_action_headers = ["Duration [s]", "Action", *list(pars_time_log.keys()), "Timestamp"]

    existing_action_headers = []
    col = 1
    while ws_action.cell(row=header_row, column=col).value is not None:
        existing_action_headers.append(ws_action.cell(row=header_row, column=col).value)
        col += 1

    union_action_headers = list(dict.fromkeys(existing_action_headers + current_call_action_headers))

    title_cell_action = ws_action.cell(row=1, column=1, value=action_sheet_name)
    title_cell_action.font = Font(bold=True, size=12, color="000000")

    if existing_action_headers != union_action_headers:
        for col_idx, header_name in enumerate(union_action_headers):
            cell = ws_action.cell(row=header_row, column=1 + col_idx, value=header_name)
            _apply_header_style(cell, action_header_color_idx)

    current_action_log_row = get_next_empty_row_in_column(ws_action, 1, header_row + 1)

    base_action_data_row_dict = {
        "Duration [s]": execution_duration,
        "Action": action,
        "Timestamp": current_timestamp,
    }
    action_data_row_dict = base_action_data_row_dict | pars_time_log

    for col_idx, header_name in enumerate(union_action_headers):
        value = action_data_row_dict.get(header_name)
        cell = ws_action.cell(row=current_action_log_row, column=1 + col_idx, value=value)
        _apply_data_style(cell, current_action_log_row, header_row)

    _set_column_widths(ws_action, union_action_headers, 1, current_action_log_row, header_row)
