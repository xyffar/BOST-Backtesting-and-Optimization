# config.py

import datetime as dt
from copy import deepcopy
from os.path import dirname, join

import streamlit as st
import yaml
from openpyxl.styles import Border, Font, PatternFill, Side

streamlit_obj = type[st.delta_generator.DeltaGenerator]
ss = st.session_state

session_state_names = {  # The first value indicate the default value when the app starts. The secondi indicates if it has to reset when a new backtest/optimization starts.
    "tickers": ([], False),  # List of tickers: list[str]
    "mode": ("backtest", False),  # Either backtest or optimization: str
    "all_strategies": ([], False),  # Dictionary of all found strategies: dict[str, CommonStrategy]
    "successful_downloads_tickers": ([], True),  # List of correctly downloaded tickers: list[str]
    "failed_downloads_tickers": ([], True),  # List of failed downloaded tickers: list[str]
    "successful_runs_tickers": ([], True),  # List of correctly backtested or optimized tickers: list[str]
    "failed_runs_tickers": ([], True),  # List of failed backtested or optimized tickers: list[str]
    "bt_params": ({}, False),  # Dictionary of params for backtest: dict[str, int|float|str]
    "backtest_results_generated": (False, True),  # Bool to indicate the presence of backtest results: bool
    "bt_stats": ({}, True),  # Dictionary of backtest result statistics: dict[str, pd.Series]
    "backtest_comp_with_benchmark_df": (
        {},
        True,
    ),  # Dictionary containing the dataframe of backtest statistics compared with benchmark: dict[str, pd.DataFrame]
    "backtest_interactive_plot": (
        {},
        True,
    ),  # Dictionary containing the interactive plots of backtest: dict[str, bokeh_plot]
    "backtest_trade_list": ({}, True),  # Dictionary of trade lists per ticker: dict[str, list]
    "mc_pars": (
        {},
        True,
    ),  # Dictionary with MC settings "# Simulations", Length Simulations and Sampling Method
    "backtest_mc_percentiles": ({}, True),
    "backtest_mc_probs_benchmark": ({}, True),
    "backtest_mc_equity_lines_plot": ({}, True),
    "matrice_equity_lines_simulati": ({}, True),  # Helper session state
    "mc_metrics_data": ({}, True),  # Helper session state containing the metrics for all the n_sim simulations
    "backtest_mc_var_plot": ({}, True),
    "backtest_mc_returns_plot": ({}, True),
    "backtest_excel": ({}, True),
    "opt_results_generated": (False, True),  # Bool to indicate the presence of optimization results: bool
    "opt_mc_results": ({}, True),
    "opt_heatmaps": ({}, True),
    "opt_sambo_plots": ({}, True),
    "opt_combs_ranking": ({}, True),
}


def initialize_session_states() -> None:
    ss.update(
        {name: deepcopy(session_state_names[name][0]) for name in session_state_names if name not in st.session_state}
    )
    # for name in session_state_names:
    #     if name not in st.session_state:
    #         ss[name] = session_state_names[name][0]


# Default dates (last 5 years until today)
# Formatted as dd/mm/yyyy
DEFAULT_START_DATE = dt.datetime(dt.datetime.now().year - 5, 1, 1).strftime("%d/%m/%Y")
MIN_DATE = dt.datetime(dt.datetime.now().year - 50, 1, 1).strftime("%d/%m/%Y")
DEFAULT_END_DATE = dt.date.today().strftime("%d/%m/%Y")

# Settings for Excel time logs
# Define a palette of distinct colors for table headers
HEADER_COLORS = [
    PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Light Blue
    PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light Green
    PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid"),  # Peach
    PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid"),  # Pink
    PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # Light Gray
    PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),  # Light Yellow
    PatternFill(start_color="E0BBE4", end_color="E0BBE4", fill_type="solid"),  # Lavender
    PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),  # Light Pink
    PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid"),  # Navajo White
    PatternFill(start_color="AFEEEE", end_color="AFEEEE", fill_type="solid"),  # Pale Turquoise
]

# Row striping colors (lighter variations or just white/light gray)
ROW_STRIPE_COLORS = [
    PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),  # White
    PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),  # Lighter Gray
]

# Font for headers
HEADER_FONT = Font(bold=True, color="000000")  # Black text

# Border style (light grey thin border)
THIN_BORDER = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)


def load_messages_from_yaml(file_path: str) -> dict:
    """Carica i messaggi testuali da un file YAML specificato."""
    try:
        with open(file_path, encoding="utf-8") as f:
            messages = yaml.safe_load(f)
            return messages if messages is not None else {}
    except FileNotFoundError:
        print(f"Errore: Il file dei messaggi YAML non trovato a '{file_path}'.")
        return {}
    except yaml.YAMLError as e:
        print(f"Errore durante il parsing del file YAML '{file_path}': {e}")
        return {}
    except Exception as e:
        print(f"Si è verificato un errore inatteso durante il caricamento di '{file_path}': {e}")
        return {}


# Carica i messaggi dal file YAML
MESSAGES = load_messages_from_yaml(join(dirname(__file__), "messages.yaml"))
